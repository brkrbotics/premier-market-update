#!/usr/bin/env python3
"""
Cloud pipeline (GitHub Actions): turn the newest reports/mwYYMM.pdf into an updated
workbook + dashboard. Run by .github/workflows/update.yml, which commits the results
(updated mw-historic-data.xlsx and index.html) back to the repo. GitHub Pages then
serves the new index.html at https://brkrbotics.github.io/premier-market-update/.

Mirrors the local "A June 2026 Automation/run_monthly.py" logic (steps 1-3); the only
differences are repo-relative paths and that publishing is the workflow's job (git),
not this script's.
"""

import argparse
import csv
import json
import re
import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
REPORTS_DIR = ROOT / "reports"
CSV_DIR = ROOT / "csv"
EXTRACTOR = ROOT / "extract_trreb.py"
XLSX = ROOT / "mw-historic-data.xlsx"
TEMPLATE = ROOT / "template.html"
DASHBOARD = ROOT / "index.html"          # GitHub Pages serves this

ORDER = ["TRREB", "Halton", "Peel", "Toronto", "York Region", "Durham", "Dufferin", "Simcoe"]
REGION_TO_AREA = {
    "TRREB": "All TRREB Areas",
    "Halton": "Halton Region",
    "Peel": "Peel Region",
    "Toronto": "City of Toronto",
    "York Region": "York Region",
    "Durham": "Durham Region",
    "Dufferin": "Dufferin County",
    "Simcoe": "Simcoe County",
}
MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]
MONTH_NUM = {m: i + 1 for i, m in enumerate(MONTHS)}

COL = dict(year=1, quarter=2, month=3, sales=4, dollar=5, avgprice=6, yoy=7, median=8,
           newlist=9, active=10, snlr=11, moi=12, moitrend=13, ldom=14, pdom=15, splp=16)

CSV_SALES, CSV_DOLLAR, CSV_AVGP, CSV_MEDIAN = "Sales", "Dollar Volume", "Average Price", "Median Price"
CSV_NEWL, CSV_SNLR, CSV_ACTIVE = "New Listings", "SNLR Trend %", "Active Listings"
CSV_MOITREND, CSV_SPLP, CSV_LDOM, CSV_PDOM = "Months Inv Trend", "Avg SP/LP %", "Avg LDOM", "Avg PDOM"

FIRST_DATA_ROW = 2


# --------------------------------------------------------------------------- helpers
def num(s):
    if s is None or s == "" or s == "-":
        return None
    if isinstance(s, (int, float)):
        return s
    s = str(s).replace("$", "").replace(",", "").rstrip("%").strip()
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return None


def period_from_stem(stem):
    m = re.search(r"(\d{2})(\d{2})$", stem)
    if not m:
        raise SystemExit(f"Cannot parse YYMM from PDF name {stem!r} (expected like mw2606).")
    yy, mm = int(m.group(1)), int(m.group(2))
    if not 1 <= mm <= 12:
        raise SystemExit(f"Bad month {mm:02d} in {stem!r}.")
    return 2000 + yy, mm, MONTHS[mm - 1], f"Q{(mm - 1) // 3 + 1}"


def find_latest_pdf():
    pdfs = [p for p in REPORTS_DIR.glob("mw*.pdf") if re.search(r"mw\d{4}$", p.stem)]
    if not pdfs:
        raise SystemExit(f"No mwYYMM.pdf found in {REPORTS_DIR}")
    return max(pdfs, key=lambda p: int(re.search(r"(\d{4})$", p.stem).group(1)))


# --------------------------------------------------------------------------- step 1
def run_extract(pdf_path):
    print(f"[1/3] Extracting {pdf_path.name} -> {CSV_DIR}")
    CSV_DIR.mkdir(exist_ok=True)
    subprocess.run([sys.executable, str(EXTRACTOR), str(pdf_path), str(CSV_DIR)], check=True)
    p3 = CSV_DIR / f"{pdf_path.stem}_p3_all_trreb_areas.csv"
    if not p3.exists():
        raise SystemExit(f"Expected CSV not produced: {p3}")
    rows = list(csv.DictReader(p3.open(encoding="utf-8")))
    print(f"      page-3 rows: {len(rows)} (expected 41)")
    by_area = {r["Area"].strip(): r for r in rows}
    missing = [a for a in REGION_TO_AREA.values() if a not in by_area]
    if missing:
        raise SystemExit(f"CSV missing expected region rows: {missing}")
    return by_area


# --------------------------------------------------------------------------- step 2
def last_data_row(ws):
    r = ws.max_row
    while r >= FIRST_DATA_ROW and ws.cell(r, COL["year"]).value in (None, ""):
        r -= 1
    return r


def append_workbook(by_area, year, month_name, quarter):
    print(f"[2/3] Appending {month_name} {year} to 8 tabs of {XLSX.name}")
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    changed = False
    for region in ORDER:
        ws = wb[region]
        last = last_data_row(ws)
        if ws.cell(last, COL["year"]).value == year and ws.cell(last, COL["month"]).value == month_name:
            print(f"      {region:12s}: {month_name} {year} already present (row {last}) — skip")
            continue
        n = last + 1
        src = by_area[REGION_TO_AREA[region]]

        def put(key, value):
            c = ws.cell(n, COL[key])
            c.value = value
            c.number_format = ws.cell(last, COL[key]).number_format

        put("year", year)
        put("quarter", quarter)
        put("month", month_name)
        put("sales", num(src[CSV_SALES]))
        put("dollar", num(src[CSV_DOLLAR]))
        put("avgprice", num(src[CSV_AVGP]))
        put("yoy", f"=IFERROR(F{n}/F{n - 12}-1,\"\")")
        put("median", num(src[CSV_MEDIAN]))
        put("newlist", num(src[CSV_NEWL]))
        put("active", num(src[CSV_ACTIVE]))
        snlr = num(src[CSV_SNLR]);  put("snlr", snlr / 100 if snlr is not None else None)
        put("moi", f"=J{n}/D{n}")
        put("moitrend", num(src[CSV_MOITREND]))
        put("ldom", num(src[CSV_LDOM]))
        put("pdom", num(src[CSV_PDOM]))
        splp = num(src[CSV_SPLP]);  put("splp", splp / 100 if splp is not None else None)
        print(f"      {region:12s}: wrote row {n}  (sales={src[CSV_SALES]}, avgPrice={src[CSV_AVGP]})")
        changed = True

    if changed:
        wb.save(XLSX)
        print("      workbook saved.")
    else:
        print("      no changes (all tabs already had this month).")
    return changed


# --------------------------------------------------------------------------- step 3
def _round(v, n=4):
    return round(v, n) if isinstance(v, (int, float)) else None


def region_rows(ws):
    last = last_data_row(ws)
    g = lambda r, k: ws.cell(r, COL[k]).value
    out = []
    for r in range(FIRST_DATA_ROW, last + 1):
        year = g(r, "year")
        month = g(r, "month")
        F = g(r, "avgprice")
        Fp = ws.cell(r - 12, COL["avgprice"]).value if r - 12 >= FIRST_DATA_ROW else None
        yoy = _round(F / Fp - 1) if isinstance(F, (int, float)) and isinstance(Fp, (int, float)) and Fp else None

        Lraw = g(r, "moi")
        if isinstance(Lraw, str) and Lraw.startswith("="):
            D, J = g(r, "sales"), g(r, "active")
            avg_moi = _round(J / D) if isinstance(J, (int, float)) and D else None
        else:
            avg_moi = Lraw if isinstance(Lraw, (int, float)) else None

        out.append({
            "year": year, "month": month, "sales": g(r, "sales"), "avgPrice": F, "yoy": yoy,
            "median": g(r, "median"), "newList": g(r, "newlist"), "activeList": g(r, "active"),
            "snlr": g(r, "snlr"), "avgMOI": avg_moi, "moiTrend": g(r, "moitrend"),
            "ldom": g(r, "ldom"), "pdom": g(r, "pdom"), "splp": g(r, "splp"),
            "label": f"{year}-{MONTH_NUM.get(month, 0):02d}",
        })
    return out


def generate_dashboard():
    print(f"[3/3] Regenerating {DASHBOARD.name} from {TEMPLATE.name}")
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    data = {"order": ORDER, "regions": {region: region_rows(wb[region]) for region in ORDER}}
    template = TEMPLATE.read_text(encoding="utf-8")
    if "{{DATA}}" not in template:
        raise SystemExit("template.html has no {{DATA}} placeholder.")
    blob = json.dumps(data, separators=(",", ":"), ensure_ascii=False)
    DASHBOARD.write_text(template.replace("{{DATA}}", blob), encoding="utf-8")
    latest = data["regions"]["TRREB"][-1]
    print(f"      latest TRREB: {latest['label']}  avgPrice={latest['avgPrice']} "
          f"yoy={latest['yoy']} avgMOI={latest['avgMOI']}")
    print(f"      wrote {DASHBOARD.name} ({DASHBOARD.stat().st_size // 1024} KB)")


# --------------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="Cloud market-stats pipeline (extract -> append -> dashboard).")
    ap.add_argument("pdf", nargs="?", help="PDF name in reports/ (default: newest mwYYMM.pdf).")
    args = ap.parse_args()

    if args.pdf:
        pdf_path = REPORTS_DIR / args.pdf if not Path(args.pdf).is_absolute() else Path(args.pdf)
    else:
        pdf_path = find_latest_pdf()
    if not pdf_path.exists():
        raise SystemExit(f"PDF not found: {pdf_path}")

    year, mm, month_name, quarter = period_from_stem(pdf_path.stem)
    print(f"Report: {pdf_path.name}  ->  {month_name} {year} ({quarter})\n")

    by_area = run_extract(pdf_path)
    append_workbook(by_area, year, month_name, quarter)
    generate_dashboard()
    print("\nPipeline done — workflow will commit any changes.")


if __name__ == "__main__":
    main()
